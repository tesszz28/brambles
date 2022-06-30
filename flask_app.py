
# A very simple Flask Hello World app for you to get started with...

from flask import Flask, request
#from rhino3dm import *
from openpyxl import load_workbook


#wb = load_workbook('C:/Users/danae/Desktop/Stats IC Workbook.xlsx')

wb = load_workbook('/home/tesszz28/brambles/Stats IC Workbook.xlsx')


app = Flask(__name__) #name of the module currently running
wsn = wb.sheetnames


@app.route('/')
def home():
    return 'Hello from UNSW CODE2120! Please type /sheets and input paramters'


@app.route('/sheets')
#def hello_page1():
#    return 'Welcome to my first non-home page, fully coded by yours truly!'
def sheetname_display():
    args = request.args
    sheetnumber = args.get("sheetnumber")
    sheetname = args.get("sheetname")
    cell = args.get("cell")
    #if sheetnumber in range(len(wb.sheetnames)): #and sheetname is None:
    if None not in (sheetnumber,sheetname):
        resp = 'You can only use one of the parameters sheetnumber & sheetname at a time'
    elif sheetnumber:
        resp = 'sheetnumber:' & wsn[int(sheetnumber)]
        if cell:
            ws = wb[resp]
            resp = ws[cell].value
    elif sheetname:
        resp = 'sheetname:' & str(wsn.index(str(sheetname)))
        if cell:
            ws = wb[sheetname]
            resp = 'cell:' & ws[cell].value
    else:
        resp = 'Please define parameters. You must pick either sheetname or sheetnumber before inputting a cell value'
    return resp

    
    # elif sheetname in list(wb.sheetnames) and sheetnumber is None:
    #     #resp_sheetnumber = str(wsn.index(str(sheetname)))
    #     ws = wb[sheetname]
    #     resp = ws[cell].value
    #else:
        #resp = 'not valid'
     #wb.sheetnames[int(sheetnumber)]
    
    #resp_sheetname = str(wsn.index(str(sheetname)))

    #result = zip(str(resp_sheetname),str(resp_sheetname))
    
    #return result
    

    #return ref
    #return wb.sheetnames[int(sheetnumber)]
    #return ref
    #sheetname = args.get("sheetname")
    #if sheetname in list(wb.sheetnames):
        #result = index(wb.sheetnames)
        #return 'There is no corresponding sheet'

    #sheetname = args.get("sheetname")
    #variables = dict(args)
    #sheets = dict(zip(range(len(wb.sheetnames)),list(wb.sheetnames)))
    
    #result = sheets

    #cells = wb.sheetnames
    # for i in range(len(wb.sheetnames)):
    #     variables[i] = wb.sheetnames[i]
    # if None not in (sheetnumber, sheetname):
    #     result = {key: value for key, value in sheets.items() if key == sheetnumber and value == sheetname}
    # elif sheetnumber is not None:
    #     result = {key: value for key, value in sheets.items() if key == sheetnumber}
    # elif sheetname is not None:
    #     result = {key: value for key, value in sheets.items() if value == sheetname}
    # #sheet_number = args.get('sheet number')
    #sheet_name = args.get('sheet name')

    #return result

# @app.route('/Statistics')
# def sheetcell_access():
#     ws = wb['Statistics']
#     return ws['A1'].value

#for i in range(len(wb.sheetnames)):
    #print(wb.sheetnames[i])
# @app.route('/sheets/<int:i>')
# def sheets(i):
#     if i in range(len(wb.sheetnames)):
#         return wsn[i]
#     else:
#         return 'not a valid request'

# @app.route('/sheets/<int:i>/<letter><int:x>')  
# def   sheetcells(i,letter,x):
#     if i in range(len(wb.sheetnames)):
#         return wsn['A1']
#     else:
#         return 'not a valid cell'


    
        #ws[i] = wb.sheetpage[i]
        #return ws['A1'].value
    #args = request.args
    #variables = dict(args)

#     @app.route('/<wb.sheetnames[i]>')
#     def sheetcell_access():
#         return wb['A1'].value

# def np_addMatrix():
#     import numpy as np
#     import pandas as pd
#     args = request.args
#     variables = dict(args)
#     variables["m1"]= variables["m1"].split(",")
#     variables["m2"]= variables["m2"].split(",")

#     variables["m1"]= list(map(lambda x: float(x), variables["m1"]))
#     variables["m2"]= list(map(lambda x: float(x), variables["m2"]))

#     matrix1 = np.array(variables["m1"]).reshape((int(variable["m1_y"]), int(variables["m1_x"])))
#     matrix2 = np.array(variables["m2"]).reshape((int(variable["m2_y"]), int(variables["m2_x"])))
    
# # Actual Function
#     result = np.add(matrix1,matrix2)
    
# # Serialisation
#     data = pd.DataFrame(result)
#     data.to_csv()

#     return str(data)
    
#     return 'Working'
def main():
    return app.run()

if __name__ == '__main__':
    main()